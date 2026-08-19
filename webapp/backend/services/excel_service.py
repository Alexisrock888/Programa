from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from datetime import datetime


GREEN_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def load_excel(file_path: str):
    return load_workbook(file_path)


def save_excel(wb, output_path: str):
    wb.save(output_path)


def set_row_green(ws, row: int, max_col: int = 10):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = GREEN_FILL


def set_row_red(ws, row: int, max_col: int = 10):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = RED_FILL


def set_cells_green(ws, row: int, columns: list):
    for col in columns:
        ws.cell(row=row, column=col).fill = GREEN_FILL


def set_cells_red(ws, row: int, columns: list):
    for col in columns:
        ws.cell(row=row, column=col).fill = RED_FILL


def generate_output_filename(original_name: str, prefix: str = "result") -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(original_name).stem
    return f"{prefix}_{stem}_{timestamp}.xlsx"


def get_rows_data(ws, start_row: int = 2):
    rows = []
    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        if any(v is not None for v in row_data):
            rows.append((row, row_data))
    return rows
